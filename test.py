import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Border, Side

# Full SHIFT_DATA preserved as per original
SHIFT_DATA = [
    ("28/11", "Day Shift"),
    ("28/11", "Night Shift"),
    ("29/11", "Day Shift"),
    ("29/11", "Night Shift"),
    ("30/11", "Day Shift"),
    ("30/11", "Night Shift"),
    ("01/12", "Day Shift"),
    ("01/12", "Night Shift"),
    ("02/12", "Day Shift"),
    ("02/12", "Night Shift"),
    ("03/12", "Day Shift"),
    ("03/12", "Night Shift"),
    ("04/12", "Day Shift"),
    ("04/12", "Night Shift"),
    ("05/12", "Day Shift"),
    ("05/12", "Night Shift"),
    ("06/12", "Day Shift"),
    ("06/12", "Night Shift"),
    ("07/12", "Day Shift"),
    ("07/12", "Night Shift"),
    ("08/12", "Day Shift"),
    ("08/12", "Night Shift"),
    ("09/12", "Day Shift"),
    ("09/12", "Night Shift"),
    ]

# Function to format the header
def format_excel_header(file_path):
    # Load the workbook and the first worksheet
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Define the header style
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Apply the styles to the header row
    for cell in sheet[1]:  # Assuming the header is in the first row
        cell.fill = header_fill
        cell.font = header_font

    # Save the workbook
    workbook.save(file_path)

# Function to add borders to all cells in the sheet
def format_excel_borders(file_path):
    # Load the workbook and the first worksheet
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Define border style
    thin_dashed_border = Border(
        left=Side(border_style="dashed", color="000000"),
        right=Side(border_style="dashed", color="000000"),
        top=Side(border_style="dashed", color="000000"),
        bottom=Side(border_style="dashed", color="000000"),
    )

    # Apply the border to all cells with data
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_dashed_border

    # Save the workbook
    workbook.save(file_path)

# Function to clean activity IDs
def clean_activity_ids(df, activity_id_index):
    df.iloc[:, activity_id_index] = df.iloc[:, activity_id_index].astype(str).str.replace(" ", "")
    return df

# Function to convert cell values to float or zero
def convert_to_float_or_zero(value):
    if isinstance(value, str):
        if value.strip() == '':
            return 0.0
        if '%' in value:
            try:
                return float(value.strip('%')) / 100
            except ValueError:
                return 0.0
        try:
            return float(value)
        except ValueError:
            return 0.0
    elif pd.isna(value):
        return 0.0
    return float(value)

# Function to find maximum progress, incorporating yesterday's output
def find_max_progress(master_df, yesterday_df, discipline_dfs, activity_id_index):
    # Define the columns explicitly including 'Overall Progress'
    max_progress_columns = [
        'Activity ID', 'Activity Description', 'Location', 'WP Owner', 'Lead craft', 'Eq Tag', 'Overall Progress'
    ] + [f"{date} ({shift})" for date, shift in SHIFT_DATA]

    # Initialize the DataFrame with the correct column names
    max_progress = pd.DataFrame(columns=max_progress_columns)
    max_progress['Activity ID'] = master_df.iloc[:, activity_id_index].unique()

    progress_bar = st.progress(0)
    total_steps = len(master_df) * len(discipline_dfs) * len(yesterday_df)
    step = 0  # Track progress for each step in total_steps

    for index, row in master_df.iterrows():
        activity_id = row.iloc[activity_id_index]
        activity_description = row.iloc[2]
        location = row.iloc[6]
        wp_owner = row.iloc[7] if len(row) > 7 else 'N/A'
        lead_craft = row.iloc[8] if len(row) > 8 else 'N/A'
        eq_tag = row.iloc[46] if len(row) > 46 else 'N/A'

        max_values = [0] * len(SHIFT_DATA)
        found = False

        # Incorporate yesterday's output if it exists
        if activity_id in yesterday_df['Activity ID'].values:
            yesterday_row = yesterday_df[yesterday_df['Activity ID'] == activity_id]
            for i, shift in enumerate(max_progress_columns[7:]):
                max_values[i] = convert_to_float_or_zero(yesterday_row[shift].values[0])

        for discipline_df in discipline_dfs:
            step += 1  # Update step count for each file processed for each activity ID
            if activity_id in discipline_df.iloc[:, activity_id_index].values:
                found = True
                row_index = discipline_df[discipline_df.iloc[:, activity_id_index] == activity_id].index[0]

                for shift_idx, shift_column in enumerate(range(11, 11 + len(SHIFT_DATA))):
                    progress_value = discipline_df.iloc[row_index, shift_column]
                    progress_value = convert_to_float_or_zero(progress_value)
                    max_values[shift_idx] = max(max_values[shift_idx], progress_value)

            progress_bar.progress(step / total_steps)

        max_progress.loc[max_progress['Activity ID'] == activity_id, 'Activity Description'] = activity_description
        max_progress.loc[max_progress['Activity ID'] == activity_id, 'Location'] = location
        max_progress.loc[max_progress['Activity ID'] == activity_id, 'WP Owner'] = wp_owner
        max_progress.loc[max_progress['Activity ID'] == activity_id, 'Lead craft'] = lead_craft
        max_progress.loc[max_progress['Activity ID'] == activity_id, 'Eq Tag'] = eq_tag
        
        # Sum the shift data and store it in 'Overall Progress' (column 6)
        max_progress.loc[max_progress['Activity ID'] == activity_id, 'Overall Progress'] = sum(max_values)

        max_progress.loc[max_progress['Activity ID'] == activity_id, max_progress_columns[7:]] = max_values

        if not found:
            st.warning(f"Activity ID '{activity_id}' not found in discipline sheets.")

    progress_bar.progress(1.0)
    return max_progress

# Streamlit app
def main():
    st.title("TA Progress Tracker - Hasdrubal TA 24")

    activity_id_index = 1  # Activity ID column index

    # Upload Master Progress Sheet
    master_file = st.file_uploader("Upload Master Progress Sheet", type=["xlsx"])
    if master_file:
        master_df = pd.read_excel(master_file, header=8)
        master_df = clean_activity_ids(master_df, activity_id_index)

        # Upload Yesterday’s Output File
        yesterday_file = st.file_uploader("Upload last day's Output File (3ajnet Emes)", type=["xlsx"])
        if yesterday_file:
            yesterday_df = pd.read_excel(yesterday_file)

            # Upload Discipline Progress Sheets
            discipline_files = st.file_uploader("Upload Discipline Progress Sheets", type=["xlsx"], accept_multiple_files=True)
            discipline_dfs = []
            for discipline_file in discipline_files:
                discipline_df = pd.read_excel(discipline_file, header=8)
                discipline_df = clean_activity_ids(discipline_df, activity_id_index)
                discipline_dfs.append(discipline_df)

            if st.button("Process Data : E3jeeeen"):
                # Ensure all required inputs are uploaded
                if not discipline_dfs:
                    st.error("Please upload at least one discipline progress sheet.")
                    return

                # Process the data
                max_progress = find_max_progress(master_df, yesterday_df, discipline_dfs, activity_id_index)

                # Save the output
                output_file = "max_progress_3ajna_output.xlsx"
                max_progress.to_excel(output_file, index=False)

                # Format the header
                format_excel_header(output_file)

                # Add borders
                format_excel_borders(output_file)

                # Provide a download link
                with open(output_file, "rb") as f:
                    st.download_button("Download the 3ajna File", data=f, file_name=output_file, mime="application/vnd.ms-excel")

if __name__ == "__main__":
    main()
